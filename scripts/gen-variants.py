# -*- coding: utf-8 -*-
"""
Genera variantes del MISMO MTO con formatos de fichero distintos.

Las 15 filas lógicas no cambian: lo que cambia es la forma del Excel, que es la dimensión que el
enunciado señala ("distintas columnas, distintos atributos, distintas abreviaturas, distinto
idioma") y la que el gold set no cubre en absoluto, porque es un único fichero.

Cada variante ataca una heurística concreta de src/pipeline/ingest.ts.
"""
import openpyxl, os, json

BASE = [
 (1,'STUD BOLT 7/8" X 130 LG, ASTM A193, GR B7 W/2 HEX. NUT 7/8", ASTM A194, GR 2H, 2 WASHER 7/8", ASTM F436','ASTM A193 GR B7/A194 GR 2H','7/8" X 130',40),
 (2,'BOLT DIN931 M20x90 with NUT DIN934 M20','A4-70','M20x90',160),
 (3,'Tornillo hexagonal DIN 933 M12 x 50 con tuerca y arandela','A2','M12x50',80),
 (4,'BOLT DIN933 M16x60 with NUT DIN934 and WASHER DIN125, 8.8, zinc plated','8.8','M16x60',100),
 (5,'STUD BOLT 1" X 150 LG, ASTM A193, GR B7, W/ 2 NUT ASTM A194, GR 2H, 1 WASHER ASTM F436','ASTM A193 GR B7','1" X 150',24),
 (6,'Tornillo DIN 931 M16 x 80 con tuerca DIN 934, 8.8, zincado','8.8','M16x80',60),
 (7,'BOLT DIN931 M12x60 A4-70 with NUT DIN934 M12 A4-80','A4-70','M12x60',50),
 (8,'HEX BOLT M16 x 70 c/w NUT AND WASHER, 8.8, ZN','8.8','M16x70',75),
 (9,'Conjunto esparrago M20 x 200 DIN 975 con 2 tuercas DIN 934 y 2 arandelas DIN 125, 8.8, zincado','8.8','M20x200',30),
 (10,'Tornillo hexagonal DIN 933 M10 x 40, 8.8, zincado','8.8','M10x40',500),
 (11,'Tuerca hexagonal DIN 934 M16, A4-80','A4-80','M16',200),
 (12,'STUD BOLT 3/4" X 110 LG, ASTM A193, GR B7','ASTM A193 GR B7','3/4" X 110',40),
 (13,'Tuerca autoblocante DIN 985 M12, 8.8, zincada','8.8','M12',300),
 (14,'Arandela plana DIN 125 M10, acero, zincada','acero','M10',250),
 (15,'Tornillo Allen cilindrico DIN 912 M10 x 40, 12.9, geomet','12.9','M10x40',120),
]
OUT='data/variants'
os.makedirs(OUT, exist_ok=True)
manifest=[]

def save(name, ataca, sheets, expect_qty=True):
    wb=openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets:
        ws=wb.create_sheet(title)
        for r in rows: ws.append(r)
    wb.save(f'{OUT}/{name}.xlsx')
    manifest.append({'file':f'{name}.xlsx','ataca':ataca,'expectQuantity':expect_qty})

# V1 control: idéntico al MTO dado
save('v01-control','ninguna (control)', [('MTO',
 [['MTO DE PRUEBA - SETS DE TORNILLERIA'],[],[],
  ['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
 [[i,d,m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V2 cabeceras en inglés y otro orden: la cantidad NO es la última columna numérica
save('v02-ingles-otro-orden','orden de columnas y detección de cantidad', [('Sheet1',
 [['ITEM','QTY','UNIT','DESCRIPTION','MATERIAL SPEC','SIZE','WEIGHT KG']] +
 [[i,c,'pcs',d,m,me,round(c*0.12,2)] for i,d,m,me,c in BASE])])

# V3 Q'TY y sin columna ITEM. expectQuantity=True: el plegado de cabeceras quita la puntuación,
# así que Q'TY sí se reconoce (hay test dedicado en variants.test.ts).
save("v03-qty-apostrofo","regex de cantidad (Q'TY) y ausencia de ITEM", [('MTO',
 [['DESCRIPTION',"Q'TY",'UOM','MATL']] +
 [[d,c,'EA',m] for i,d,m,me,c in BASE])])

# V4 sin columna de cantidad en absoluto
save('v04-sin-cantidad','ausencia total de cantidad', [('BOM',
 [['POS','DESCRIPCION','MATERIAL','MEDIDA']] +
 [[i,d,m,me] for i,d,m,me,c in BASE])], expect_qty=False)

# V5 descripción partida en dos columnas
save('v05-descripcion-partida','concatenación de todas las celdas de texto', [('MTO',
 [['ITEM','DESCRIPCION','OBSERVACIONES','MATERIAL','MEDIDA','CANT.','UD']] +
 [[i, d.split(',')[0], ','.join(d.split(',')[1:]).strip(), m, me, c,'uds'] for i,d,m,me,c in BASE])])

# V6 columnas de ruido de proyecto, cantidad en medio
save('v06-columnas-ruido','columnas irrelevantes y cantidad no final', [('MTO',
 [['PROYECTO','WBS','REV','ITEM','DESCRIPCION','CANTIDAD','UD','MATERIAL','MEDIDA','PESO']] +
 [['P-2291','2291-PIP-03','C',i,d,c,'uds',m,me,round(c*0.31,2)] for i,d,m,me,c in BASE])])

# V7 bloque de título largo y una columna en blanco por delante
save('v07-titulo-largo','detección de la fila de cabeceras', [('MTO',
 [['CLIENTE: EPC ASTURIANA S.A.'],['OBRA: PLANTA DE TRATAMIENTO'],
  ['DOC: 2291-MTO-TOR-012'],['REV: 12'],['FECHA: 2026-08-01'],[],[],
  [None,'ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
 [[None,i,d,m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V8 los datos están en la SEGUNDA hoja; la primera es una portada
save('v08-segunda-hoja','selección de hoja', [
 ('PORTADA',[['MEMORIA DE MATERIALES'],['Rev 12'],['Ingeniería subcontratada']]),
 ('TORNILLERIA',[['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
  [[i,d,m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V9 tipos sucios: cantidades como texto con unidad pegada, números como cadena
save('v09-tipos-sucios','lectura de celdas con tipos inconsistentes', [('MTO',
 [['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
 [[str(i),d,m,me,(f'{c}' if i%2 else f'{c},00'),'uds'] for i,d,m,me,c in BASE])])

# V10 cabeceras en francés (estudio externo). expectQuantity=True: QUANTITÉ se pliega a
# QUANTITE, que sí es token de cantidad.
save('v10-frances','cabeceras en otro idioma', [('Feuille1',
 [['REPÈRE','DÉSIGNATION','MATIÈRE','DIMENSION','QUANTITÉ','UNITÉ']] +
 [[i,d,m,me,c,'pcs'] for i,d,m,me,c in BASE])])

# V11 filas en blanco intercaladas: se descartan como EMPTY_ROW y se reportan, no se omiten en silencio
rows=[]
for k,(i,d,m,me,c) in enumerate(BASE):
    rows.append([i,d,m,me,c,'uds'])
    if k in (2,7,12): rows.append([''])
save('v11-blancos-intercalados','filas en blanco intercaladas (EMPTY_ROW reportado)', [('MTO',
 [['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] + rows)])

# V12 ITEM alfanumérico: la referencia no siempre es un número
save('v12-item-alfanumerico','referencias de línea alfanuméricas (TOR-001)', [('MTO',
 [['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
 [[f'TOR-{i:03d}',d,m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V13 tres hojas: portada + DOS hojas de datos; sólo la primera con cabeceras se procesa
save('v13-tres-hojas','dos hojas con datos: la segunda se ignora y se reporta', [
 ('PORTADA',[['MEMORIA DE MATERIALES'],['Rev 12']]),
 ('TORNILLERIA_A',[['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
  [[i,d,m,me,c,'uds'] for i,d,m,me,c in BASE]),
 ('TORNILLERIA_B',[['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
  [[i,d,m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V14 primera hoja completamente vacía; los datos están en la segunda
save('v14-hoja-vacia','primera hoja vacía, datos en la segunda', [
 ('VACIA',[]),
 ('MTO',[['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
  [[i,d,m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V15 carátula de 24 filas: la cabecera cae en la fila 25, justo en el límite del escaneo
save('v15-cabecera-en-el-limite','detección de cabeceras en el límite del escaneo (fila 25)', [('MTO',
 [['PROYECTO: P-2291'],['CLIENTE: EPC ASTURIANA'],['DOC: 2291-MTO-TOR-012'],['REV: C'],
  ['FECHA: 2026-08-01'],['ELABORADO: DEPT. MATERIALES'],['REVISADO:'],['APROBADO:'],
  [],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],
  ['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
 [[i,d,m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V16 cabecera con salto de línea: el plegado de cabeceras la reconoce
save('v16-salto-linea-cabecera',"cabecera con salto de línea ('CANT.\\nTOTAL')", [('MTO',
 [['ITEM','DESCRIPCIÓN','MATERIAL','MEDIDA','CANT.\nTOTAL','UD']] +
 [[i,d,m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V17 cabeceras en minúsculas y con diacríticos
save('v17-minusculas-diacriticos','cabeceras en minúsculas con diacríticos', [('MTO',
 [['ítem','descripción','material','medida','cantidad','ud']] +
 [[i,d,m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V18 cabeceras en alemán: MENGE no es un token de cantidad reconocido
save('v18-aleman','cabeceras en alemán: MENGE no reconocida, se avisa con candidatas', [('Blatt1',
 [['POS','BEZEICHNUNG','WERKSTOFF','ABMESSUNG','MENGE','EINHEIT']] +
 [[i,d,m,me,c,'Stück'] for i,d,m,me,c in BASE])], expect_qty=False)

# V19 cabeceras en portugués: QTDE tampoco es token reconocido
save('v19-portugues','cabeceras en portugués: QTDE no reconocida, se avisa', [('Folha1',
 [['ITEM','DESCRIÇÃO','MATERIAL','MEDIDA','QTDE','UD']] +
 [[i,d,m,me,c,'pçs'] for i,d,m,me,c in BASE])], expect_qty=False)

# V20 cantidad como fórmula sin valor cacheado: no se inventa, sale null
save('v20-formulas-sin-cache','cantidad como fórmula sin valor cacheado (null, no inventada)', [('MTO',
 [['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
 [[i,d,m,me,f'={c}','uds'] for i,d,m,me,c in BASE])], expect_qty=False)

# V21 la única columna de cantidad se llama UDS: token de último recurso, y no hay unidad aparte
save('v21-uds-como-cantidad',"cantidad en columna 'UDS' (token de último recurso)", [('MTO',
 [['ITEM','DESCRIPCION','MATERIAL','MEDIDA','UDS']] +
 [[i,d,m,me,c] for i,d,m,me,c in BASE])])

# V22 NOS como cantidad y UOM como unidad (jerga americana)
save('v22-nos-como-cantidad',"cantidad 'NOS' y unidad 'UOM'", [('MTO',
 [['ITEM','DESCRIPTION','MATL','SIZE','NOS','UOM']] +
 [[i,d,m,me,c,'EA'] for i,d,m,me,c in BASE])])

# V23 QTY con pesos y CANT. con las cantidades: la prioridad de la tabla manda, no el orden
save('v23-prioridad-cant-sobre-qty',"dos columnas candidatas: 'CANT.' gana a 'QTY' aunque vaya después", [('MTO',
 [['ITEM','DESCRIPCION','QTY','MATERIAL','MEDIDA','CANT.','UD']] +
 [[i,d,round(c*0.12,2),m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V24 dos columnas llamadas MATERIAL: la calidad está en la segunda, como en el MTO real
save('v24-cabecera-duplicada',"cabecera duplicada ('MATERIAL' dos veces)", [('MTO',
 [['ITEM','DESCRIPCION','MATERIAL','MATERIAL','MEDIDA','CANT.','UD']] +
 [[i,d,'acero',m,me,c,'uds'] for i,d,m,me,c in BASE])])

# V25 título con celdas combinadas sobre las cabeceras. El merge se limita a 2 columnas a
# propósito: ExcelJS replica el valor maestro en todas las celdas del rango, así que un título
# combinado sobre ≥3 columnas simula una fila de cabeceras y rompe la detección. Ese hueco
# (banner combinado ancho) queda documentado en docs/10-benchmarks.md.
wb=openpyxl.Workbook(); wb.remove(wb.active)
ws=wb.create_sheet('MTO')
ws.append(['MTO TORNILLERÍA — PLANTA DE TRATAMIENTO']); ws.merge_cells('A1:B1')
ws.append([]); ws.append([])
ws.append(['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD'])
for i,d,m,me,c in BASE: ws.append([i,d,m,me,c,'uds'])
wb.save(f'{OUT}/v25-celdas-combinadas.xlsx')
manifest.append({'file':'v25-celdas-combinadas.xlsx','ataca':'celdas combinadas sobre la cabecera','expectQuantity':True})

# V26 columna de fechas reales (objetos Date, no texto)
import datetime
save('v26-fechas','celdas de fecha (Date) junto a los datos', [('MTO',
 [['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD','ENTREGA']] +
 [[i,d,m,me,c,'uds',datetime.date(2026,9,(i%28)+1)] for i,d,m,me,c in BASE])])

# V27 cantidades como texto con espacios duros (copia desde PDF)
save('v27-espacios-duros','cantidades con espacios duros (NBSP) alrededor', [('MTO',
 [['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
 [[i,d,m,me,f' {c} ','uds'] for i,d,m,me,c in BASE])])

# V28 una celda extra sin cabecera al final de cada fila: la anchura la fija la cabecera
save('v28-celda-huerfana','celdas más allá de la última cabecera se descartan', [('MTO',
 [['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.','UD']] +
 [[i,d,m,me,c,'uds','nota sin cabecera'] for i,d,m,me,c in BASE])])

# V29 columnas en orden inverso: nada depende de la posición
save('v29-orden-inverso','columnas en orden inverso', [('MTO',
 [['UD','CANT.','MEDIDA','MATERIAL','DESCRIPCION','ITEM']] +
 [['uds',c,me,m,d,i] for i,d,m,me,c in BASE])])

# V30 unidad pegada a la cantidad en la misma celda
save('v30-unidad-pegada',"unidad pegada a la cantidad ('40 uds')", [('MTO',
 [['ITEM','DESCRIPCION','MATERIAL','MEDIDA','CANT.']] +
 [[i,d,m,me,f'{c} uds'] for i,d,m,me,c in BASE])])

json.dump(manifest, open(f'{OUT}/manifest.json','w'), indent=2, ensure_ascii=False)
print(f'{len(manifest)} variantes en {OUT}/')
for m in manifest: print(f"  {m['file']:28} ataca: {m['ataca']}")
